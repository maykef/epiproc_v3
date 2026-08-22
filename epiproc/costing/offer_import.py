"""Bulk import of a W. Tuning "offer" spreadsheet into products + draft costings.

Three layers, matching the phase-1 split (pure arithmetic in calc.py, DB access
behind epiproc.db.costing):

  ``parse_offer``   — pure parser (openpyxl only, no I/O beyond the given file).
                      One :class:`OfferRow` per product row; every coercion
                      problem becomes a warning, never an exception. The file is
                      hand-typed and messy, so nothing here hard-codes the row
                      layout — the data region is *detected*.

  ``build_inputs``  — pure mapping of one OfferRow into the SAME
                      ``calc.CostingInputs`` model the phase-1 calculator uses.
                      Per-product values (material cost N, UPT J, box by height
                      L) come from the row; everything shared (rates, inbound/
                      outbound constants, the default menu selection) comes from
                      ``costing_defaults`` + the seeded menus/box prices. Returns
                      None when the row can't be costed yet (missing UPT /
                      material cost / resolvable box) — the row still imports as
                      a product, it just gets no costing.

  ``run_offer_import`` — batch driver. Non-dry runs happen in ONE transaction
                      (a mid-batch failure rolls everything back); dry_run
                      computes and reports identically but writes nothing, and
                      backs the admin preview screen.

Import invariants: never auto-finalise (every imported costing is a draft for
review); re-importing the same offer updates the same products and adds a NEW
draft version (never duplicate products); an update never clobbers
selling/retail an admin set; a product with no stored selling price gets one
auto-targeted from ``our_target_margin`` (the workbook's Target FP formula,
rounded to 2dp) so every row shows cost → profit → total immediately; each
confirmed upload also writes one ``offer_imports`` batch row (one upload = one
version); column M is deliberately ignored — the material cost is column N
("in box, on pallet, with sticker"), per the operator.
"""
from __future__ import annotations

import io
import re
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
from pydantic import BaseModel

from epiproc.costing.calc import (
    BoxSelection,
    CostingInputs,
    InboundTransport,
    MaterialLine,
    MenuSelection,
    Outbound,
    PackagingPcts,
    compute,
)

# ── OfferRow ──────────────────────────────────────────────────────────────────


class OfferRow(BaseModel):
    row_number: int
    name: str
    ean: str | None = None
    upt: int | None = None               # col J — pcs per tray
    material_cost: Decimal | None = None  # col N — price/pc in box on pallet
    box_height_cm: int | None = None      # col L — first integer wins
    layers_per_cc: int | None = None      # col H
    trays_per_layer: int | None = None    # col I
    per_pallet_raw: str | None = None     # col K — stored for display only
    warnings: list[str] = []


# ── Parsing ───────────────────────────────────────────────────────────────────

_MONEY_PLACES = Decimal("0.0001")
_MONEY_PLACES_2 = Decimal("0.01")   # auto selling price: rounded to 2dp
_DIGITS = re.compile(r"\d+")


def _to_text(value: Any) -> str:
    """Coerce a cell to text: ints/floats without float noise, else str()."""
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def _to_decimal(value: Any) -> Decimal | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, Decimal):
        return value
    text = _to_text(value)
    if not text:
        return None
    try:
        # Quantize so float artefacts (1.8524999999999998 -> 1.8525) don't leak
        # into saved snapshots; the offer prices are 4-dp.
        return Decimal(text).quantize(_MONEY_PLACES, rounding=ROUND_HALF_UP)
    except InvalidOperation:
        return None


def _to_int(value: Any) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    text = _to_text(value)
    if not text.isdigit():
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _first_int(text: str) -> int | None:
    m = _DIGITS.search(text)
    return int(m.group()) if m else None


def _warn(row: dict, msg: str) -> None:
    if msg not in row["warnings"]:
        row["warnings"].append(msg)


def parse_offer(path_or_bytes: str | bytes | Path) -> list[OfferRow]:
    """Parse an offer workbook into OfferRows. Defensive by design: the data
    region starts at the first row whose col B is a non-empty name AND col J or
    N is numeric (the header block above is skipped without hard-coding row 4),
    and no bad cell ever raises — it becomes a warning on the row.

    Column semantics (letters as in the file):
      B name · C EAN/GTIN · H layers per cc · I trays per layer ·
      J pcs per tray (UPT) · K per pallet (kept raw) · L box height ·
      N price/pc in box on pallet (material cost). Column M is ignored.
    """
    if isinstance(path_or_bytes, (str, Path)):
        wb = openpyxl.load_workbook(path_or_bytes, read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(path_or_bytes), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        rows: list[OfferRow] = []
        started = False
        for row_idx, values in enumerate(ws.iter_rows(values_only=True), start=1):
            # values is a tuple; columns beyond the sheet width are absent.
            def cell(idx: int) -> Any:
                return values[idx] if idx < len(values) else None

            b = _to_text(cell(1))
            j_raw, n_raw = cell(9), cell(13)
            if not started:
                # Header block: B holds a label ("NAME", supplier, title) and
                # the data columns aren't numeric. The first row where B is a
                # real name AND J or N is numeric opens the data region.
                if b and (_to_int(j_raw) is not None or _to_decimal(n_raw) is not None):
                    started = True
                else:
                    continue

            if not b:
                # An empty-name row inside the data region is skipped, but only
                # worth reporting if something else is in it (truly empty
                # trailing rows are not product lines at all).
                if any(cell(i) not in (None, "") for i in range(20)):
                    rows.append(OfferRow(row_number=row_idx, name="",
                                         warnings=["empty_name"]))
                continue

            row: dict = {"row_number": row_idx, "name": b, "warnings": []}

            ean = _to_text(cell(2))
            if ean.upper() == "#VALUE!":
                ean = ""
            if ean and not ean.isdigit():
                ean = ""                       # non-digit junk -> None
            if ean:
                if len(ean) in (8, 13):
                    row["ean"] = ean
                else:
                    row["ean"] = ean
                    _warn(row, "ean_suspect")  # digits but wrong length: keep, flag

            upt = _to_int(cell(9))
            if upt is None or upt <= 0:
                _warn(row, "upt_missing")
            else:
                row["upt"] = upt

            cost = _to_decimal(cell(13))
            if cost is None or cost <= 0:
                _warn(row, "material_cost_missing")
            else:
                row["material_cost"] = cost

            l_raw = cell(11)
            if l_raw is not None:
                l_text = _to_text(l_raw)
                box_h = _to_int(l_raw) if l_text.isdigit() else _first_int(l_text)
                if box_h is not None:
                    row["box_height_cm"] = box_h
                else:
                    _warn(row, "box_unresolved")
            else:
                _warn(row, "box_unresolved")

            h, i = _to_int(cell(7)), _to_int(cell(8))
            if h is not None:
                row["layers_per_cc"] = h
            if i is not None:
                row["trays_per_layer"] = i

            k_raw = cell(10)
            if k_raw is not None and _to_text(k_raw):
                row["per_pallet_raw"] = _to_text(k_raw)

            rows.append(OfferRow(**row))
        return rows
    finally:
        wb.close()


# ── Mapping: OfferRow -> CostingInputs ────────────────────────────────────────

# The default menu selection every imported product gets (roses costing values):
# Consumables (per-case) on, everything else off — no sleeve logic this phase.
# Equipment is off; labour has Pack on line + Labelling on.
_DEFAULT_ON_PACKAGING = {"consumables"}
_DEFAULT_ON_LABOUR = {"pack on line", "labelling"}
# Same customer-specific glue as routers/costing.py (_PER_PACK_EQUIPMENT): the
# palletised-equipment divisor is keyed by name, kept out of the pure calc.
_DIVIDE_BY_PACK = {"chep pallets"}

# Box dimensions look like "60х40х34" (the separator varies: Cyrillic х, ASCII
# x, ×, *). The LAST component is the height in cm.
_BOX_SEP = re.compile(r"[хxX×*]+")


def resolve_box(box_height_cm: int, box_types: list[dict]) -> dict | None:
    """Smallest real box model the requested height fits into (height >= the
    row's height). The operator confirmed (2026-08-19) that only four box
    models exist — 34/40/48/80 cm — so every offer height must land in one of
    those; a height taller than every model resolves to None."""
    fitting = []
    for bt in box_types:
        dims = bt.get("dimensions")
        if not dims:
            continue
        parts = [p for p in _BOX_SEP.split(str(dims)) if p]
        if not parts:
            continue
        height = _to_int(parts[-1])
        if height is not None and height >= box_height_cm:
            fitting.append((height, bt))
    if not fitting:
        return None
    fitting.sort(key=lambda pair: pair[0])
    return fitting[0][1]


def build_inputs(
    row: OfferRow,
    defaults: dict,
    box_types: list[dict],
    menu_items: list[dict],
) -> CostingInputs | None:
    """Map one OfferRow onto the phase-1 CostingInputs. Returns None when the
    row can't be costed yet (missing UPT / material cost / box height / fitting
    box type) — the caller still imports the product, just with no costing.
    A failed box resolution appends ``box_unresolved`` to row.warnings so the
    reason reaches the review table."""
    if row.upt is None or row.material_cost is None:
        return None
    if row.box_height_cm is None:
        return None
    box_type = resolve_box(row.box_height_cm, box_types)
    if box_type is None:
        if "box_unresolved" not in row.warnings:
            row.warnings.append("box_unresolved")
        return None

    def sel(items: list[dict], kind: str) -> list[MenuSelection]:
        out = []
        for it in items:
            if it.get("kind") != kind:
                continue
            name = str(it["name"])
            folded = name.casefold()
            qty = 0.0
            if kind in ("packaging_per_unit", "packaging_per_case"):
                qty = 1.0 if folded in _DEFAULT_ON_PACKAGING else 0.0
            elif kind == "labour":
                qty = 1.0 if folded in _DEFAULT_ON_LABOUR else 0.0
            out.append(MenuSelection(
                name=name,
                unit_cost=float(it["unit_cost"]),
                qty=qty,
                kind=kind,
                divide_by_pack=(kind == "equipment" and folded in _DIVIDE_BY_PACK),
            ))
        return out

    def d(key: str, fallback: float = 0.0) -> float:
        v = defaults.get(key)
        return float(v) if v is not None else fallback

    return CostingInputs(
        units_per_tray=row.upt,
        eur_rate=d("eur_rate", 1.0),
        materials=[MaterialLine(name=row.name, unit_cost=float(row.material_cost),
                                qty=1.0)],
        inbound=InboundTransport(
            qty_per_box=d("qty_per_box", 1.0),
            boxes_per_cc=d("boxes_per_cc"),
            pallet_rate=d("pallet_rate"),
        ),
        waste_pct=d("waste_pct"),
        packaging=sel(menu_items, "packaging_per_unit") + sel(menu_items, "packaging_per_case"),
        packaging_pct=PackagingPcts(),          # all zero this phase
        box=BoxSelection(code=box_type["code"], price=float(box_type["price"])),
        equipment=sel(menu_items, "equipment"),
        outbound=Outbound(
            price_per_pallet=d("price_per_pallet"),
            fill_rate=d("fill_rate", 1.0),
            boxes_on_order=d("boxes_on_order"),
            fuel_surcharge_pct=d("fuel_surcharge_pct"),
        ),
        labour=sel(menu_items, "labour"),
        intake_labour_pct=d("intake_labour_pct"),
        additional_pct=d("additional_pct"),
        vat_rate=d("vat_rate"),
        retail_price=None,                     # unknown per product on import
        selling_price=None,                    # (margins come back None — fine)
        customer_target_margin=d("customer_target_margin"),
        our_target_margin=d("our_target_margin"),
    )


# ── Batch driver ──────────────────────────────────────────────────────────────


class RowReport(BaseModel):
    row_number: int
    name: str
    ean: str | None = None
    upt: int | None = None
    material_cost: Decimal | None = None
    box_code: str | None = None
    per_pallet_raw: str | None = None
    status: str                             # "costed" | "attention" | "skipped"
    total_direct_cost: float | None = None
    selling_price: float | None = None      # stored price, else the auto target
    retail_price: float | None = None       # stored retail, else the auto target
    our_gp: float | None = None             # profit = selling - direct cost
    customer_gp_pct: float | None = None    # customer margin on OUR price
    product_action: str | None = None       # "created" | "updated"
    warnings: list[str] = []


class ImportReport(BaseModel):
    source: str = ""
    products_created: int = 0
    products_updated: int = 0
    costings_created: int = 0
    skipped: int = 0
    batch_id: int | None = None             # offer_imports row (non-dry only)
    rows: list[RowReport] = []


def _row_report(row: OfferRow, **extra: Any) -> RowReport:
    return RowReport(
        row_number=row.row_number,
        name=row.name,
        ean=row.ean,
        upt=row.upt,
        material_cost=row.material_cost,
        per_pallet_raw=row.per_pallet_raw,
        warnings=list(row.warnings),
        **extra,
    )


def run_offer_import(
    rows: list[OfferRow],
    actor: str,
    *,
    dry_run: bool = False,
    source: str = "",
    finalise: bool = False,
) -> ImportReport:
    """Import every parsed row: upsert the product (keyed per
    :func:`epiproc.db.costing.find_product_by_key`), then — when the row is
    costable — save a costing computed through the shared calc. Never
    auto-finalises: the costing is a DRAFT unless the caller passes
    ``finalise=True`` (the import wizard's explicit bulk-finalise checkbox —
    still an admin decision, never automatic). dry_run computes and reports
    identically but writes nothing; the real run wraps every write in one
    transaction."""
    from epiproc.db import costing as db
    from epiproc.db.pool import pool

    defaults = db.get_costing_defaults()
    box_types = db.list_box_types(active_only=True)
    menu_items = db.list_menu_items(active_only=True)

    report = ImportReport(source=source)
    seen_key_eans: set[str] = set()
    dry_seen: dict[str, dict] = {}  # dry run: key -> simulated stored product

    def handle(row: OfferRow, conn=None,  # noqa: ANN001
               offer_import_id: int | None = None) -> None:
        if not row.name:
            report.skipped += 1
            report.rows.append(_row_report(row, status="skipped"))
            return

        # Batch-level duplicate detection: two rows sharing a valid EAN would
        # merge into one product — surface that instead of silently overwriting.
        key_ean = row.ean if (row.ean and len(row.ean) in (8, 13)) else None
        if key_ean and key_ean in seen_key_eans:
            row.warnings.append("duplicate_ean")
        if key_ean:
            seen_key_eans.add(key_ean)
        elif row.ean and "ean_suspect" not in row.warnings:
            # Flag junk-length EANs here too (not just in the parser) so rows
            # re-validated from the confirm payload are never silently dropped.
            row.warnings.append("ean_suspect")

        # Look the product up up-front (not as a side effect of the upsert) so
        # both paths know whether a selling price is already stored — a stored
        # price always wins over the auto-target computed below.
        # Junk-length EANs are display data only: products.ean is UNIQUE, so
        # persisting a bad formula result could make two distinct products
        # collide and roll back the whole batch. They stay on the row as
        # ean_suspect warnings for the review table.
        db_ean = row.ean if (row.ean and len(row.ean) in (8, 13)) else None
        if conn is not None:
            existing = db.find_product_by_key(row.name, db_ean, conn=conn)
        else:  # dry run: read-only prediction of the same action — rows written
            # earlier in this batch are simulated via dry_seen so the counts and
            # the stored-price-wins behaviour mirror the transactional run
            # exactly (two rows sharing one EAN merge into one product there).
            key = db_ean if db_ean else row.name.strip()
            existing = dry_seen.get(key) or db.find_product_by_key(row.name, db_ean)
        created = existing is None
        # Only an admin-set ("human") price wins. A price a previous import
        # generated is stale derived data: price movements arrive in the
        # supplier's offers, so it is recomputed from this offer's cost.
        admin_priced = (existing is not None
                        and existing.get("price_origin") == "human")
        stored_selling = None
        if admin_priced and existing.get("selling_price") is not None:
            stored_selling = float(existing["selling_price"])
        stored_retail = None
        if admin_priced and existing.get("retail_price") is not None:
            stored_retail = float(existing["retail_price"])
        if created:
            report.products_created += 1
        else:
            report.products_updated += 1

        inputs = build_inputs(row, defaults, box_types, menu_items)
        if inputs is None:
            # Uncostable rows still import as products — but with no selling
            # price: the auto-target needs a direct cost to divide.
            if conn is not None:
                db.upsert_product_by_key(row.name, db_ean, row.upt, conn)
            else:
                dry_seen[key] = {"selling_price": stored_selling,
                                 "retail_price": stored_retail}
            report.rows.append(_row_report(
                row, status="attention",
                product_action="created" if created else "updated"))
            return

        results = compute(inputs)
        # Selling price: a stored one (re-imports) wins; otherwise auto-target
        # from our_target_margin — the workbook's Target FP formula — so every
        # imported row shows cost → profit → total with no manual steps. The
        # chosen price is written into the inputs snapshot too, so the saved
        # costing records exactly what it was computed against.
        selling = stored_selling
        if selling is None:
            margin = defaults.get("our_target_margin")
            if margin is not None and float(margin) < 1.0:
                selling = float(
                    (Decimal(str(results.total_direct_cost))
                     / (Decimal(1) - Decimal(str(margin))))
                    .quantize(_MONEY_PLACES_2, rounding=ROUND_HALF_UP))
            elif "margin_invalid" not in row.warnings:
                row.warnings.append("margin_invalid")
        if selling is not None:
            inputs = inputs.model_copy(update={"selling_price": selling})
            results = compute(inputs)
        # Retail: a stored one (admin-set or earlier import) wins; otherwise the
        # workbook's Target Retail — selling ÷ (1 − customer_target_margin) ×
        # (1 + vat) — so customer GP% fills in with zero manual steps too.
        retail = stored_retail
        if retail is None and selling is not None:
            cm = defaults.get("customer_target_margin")
            vat = defaults.get("vat_rate")
            if cm is not None and vat is not None and float(cm) < 1.0:
                retail = float(
                    (Decimal(str(selling)) / (Decimal(1) - Decimal(str(cm)))
                     * (Decimal(1) + Decimal(str(vat))))
                    .quantize(_MONEY_PLACES_2, rounding=ROUND_HALF_UP))
        if retail is not None:
            inputs = inputs.model_copy(update={"retail_price": retail})
            results = compute(inputs)

        if conn is not None:
            product_id, _ = db.upsert_product_by_key(
                row.name, db_ean, row.upt, conn, selling_price=selling,
                retail_price=retail)
            db.save_costing(product_id, inputs.model_dump(), results.as_floats(),
                            created_by=actor,
                            status="final" if finalise else "draft", conn=conn,
                            offer_import_id=offer_import_id)
        else:
            dry_seen[key] = {"selling_price": selling, "retail_price": retail}
        report.costings_created += 1
        report.rows.append(_row_report(
            row, status="costed",
            box_code=inputs.box.code,
            total_direct_cost=float(results.total_direct_cost),
            selling_price=selling,
            retail_price=retail,
            our_gp=float(results.our_gp) if results.our_gp is not None else None,
            customer_gp_pct=(
                float(results.customer_gp_pct)
                if results.customer_gp_pct is not None else None),
            product_action="created" if created else "updated"))

    if dry_run:
        for row in rows:
            handle(row)
        return report

    with pool().connection() as conn:        # one transaction: all-or-nothing
        # The batch row is opened first so each costing can reference it; its
        # counts are written below. A failure rolls the whole thing back, batch
        # row included, so a half-imported offer never appears in the history.
        batch_id = db.create_offer_import(source, actor, conn, finalised=finalise)
        report.batch_id = batch_id
        for row in rows:
            handle(row, conn=conn, offer_import_id=batch_id)
        db.update_offer_import_counts(
            batch_id, len(rows), report.products_created,
            report.products_updated, report.costings_created, report.skipped,
            conn=conn)
    return report
